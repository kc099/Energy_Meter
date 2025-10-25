import logging
import json
import csv
import re
from datetime import datetime, timedelta, time

# Celery is optional for local/testing environments
try:
    from celery.exceptions import CeleryError
    from kombu.exceptions import OperationalError
except ModuleNotFoundError:  # pragma: no cover - optional dependency guard
    class CeleryError(Exception):
        pass

    class OperationalError(Exception):
        pass
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Avg, Min, Max, Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.utils.dateparse import parse_datetime
from urllib.parse import urlparse

from accounts.models import User

from .forms import DeviceForm
from .models import (
    Device,
    Shift,
    ShiftReport,
    DeviceData,
    DeviceShare,
    DeviceProvisioningToken,
)
from .tasks import poll_device_task

from andon.models import Station, ShiftData, SectionData

logger = logging.getLogger(__name__)


_NUMERIC_PATTERN = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?")


def _coerce_float(value):
    """Best-effort conversion of telemetry values to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
        try:
            return float(candidate)
        except ValueError:
            match = _NUMERIC_PATTERN.search(candidate)
            if match:
                try:
                    return float(match.group(0))
                except ValueError:
                    return None
    return None


def _extract_numeric(payload: dict, primary_keys: tuple[str, ...], *, fuzzy: tuple[str, ...] = ()):
    """Pull the first numeric value from telemetry payload matching known keys."""
    if not isinstance(payload, dict):
        return None

    for key in primary_keys:
        if key in payload:
            coerced = _coerce_float(payload.get(key))
            if coerced is not None:
                return coerced

    if not fuzzy:
        return None

    for key, value in payload.items():
        lowered = key.lower()
        if any(token in lowered for token in fuzzy):
            coerced = _coerce_float(value)
            if coerced is not None:
                return coerced

    return None


def _get_device_role(device: Device, user) -> str | None:
    if not user.is_authenticated:
        return None
    if device.device_owner_id == getattr(user, 'id', None):
        return DeviceShare.AccessLevel.MANAGER
    share = device.device_shares.filter(user=user).first()
    return share.role if share else None


def _address_candidates(raw_address: str) -> list[str]:
    if not raw_address:
        return []
    addr = raw_address.strip()
    if not addr:
        return []
    candidates = [addr]

    parsed = urlparse(addr if '://' in addr else f'//{addr}', scheme='http')
    host = parsed.hostname or ''
    if host:
        if parsed.port:
            candidates.append(f'{host}:{parsed.port}')
        candidates.append(host)

    head = addr.split('/', 1)[0]
    if head:
        candidates.append(head)
        if ':' in head:
            candidates.append(head.split(':', 1)[0])

    # deduplicate while preserving order
    seen = set()
    unique = []
    for candidate in candidates:
        candidate = candidate.strip()
        if candidate and candidate not in seen:
            seen.add(candidate)
            unique.append(candidate)
    return unique


def _find_andon_station(device: Device):
    for candidate in _address_candidates(device.device_address):
        station = Station.objects.filter(ip_address__iexact=candidate).first()
        if station:
            return station
    return None


def _build_andon_snapshot(station: Station) -> dict:
    today = timezone.localdate()
    shift_entry = (
        ShiftData.objects.filter(station=station, date=today)
        .order_by('-id')
        .first()
    )
    section_entry = (
        SectionData.objects.filter(station=station).order_by('-id').first()
    )

    return {
        'id': station.id,
        'name': station.name,
        'plan': shift_entry.plan if shift_entry else station.plan_shift1,
        'actual': shift_entry.actual if shift_entry else station.actual_count,
        'downtime_min': (shift_entry.downtime_min if shift_entry else station.total_downtime_min),
        'fault_time': section_entry.fault_time if section_entry else None,
        'resolved_time': section_entry.resolved_time if section_entry else None,
        'ip': station.ip_address,
        'created_at': station.created_at,
        'last_updated': station.last_ping,
        'is_active': station.is_active and station.is_alive,
    }

def generate_shift_reports_for_date(target_date):
    """Generate shift reports for all devices and shifts for a specific date"""
    generated_count = 0
    devices = Device.objects.filter(is_active=True)
    shifts = Shift.objects.filter(is_active=True)

    for device in devices:
        for shift in shifts:
            # Check if report already exists
            # Generate report for this shift and device
            start_datetime = timezone.make_aware(datetime.combine(target_date, shift.start_time))
            end_datetime = timezone.make_aware(datetime.combine(target_date, shift.end_time))

            # If shift ends next day
            if shift.start_time > shift.end_time:
                end_datetime += timedelta(days=1)

            # Get data for the shift period
            shift_data = DeviceData.objects.filter(
                device=device,
                timestamp__range=(start_datetime, end_datetime)
            ).order_by('timestamp')

            entries = list(shift_data)
            samples: list[tuple[DeviceData, dict]] = [
                (entry, entry.value)
                for entry in entries
                if isinstance(entry.value, dict)
            ]
            if not samples:
                continue

            pf_samples: list[tuple[DeviceData, float]] = []
            current_samples: list[tuple[DeviceData, float]] = []
            voltage_samples: list[tuple[DeviceData, float]] = []
            for entry, payload in samples:
                try:
                    pf_value = _extract_numeric(
                        payload,
                        ("power_factor", "pf", "powerfactor"),
                        fuzzy=("pf", "powerfactor"),
                    )
                    if pf_value is None:
                        raise ValueError
                    pf_samples.append((entry, pf_value))
                except ValueError:
                    continue

                current_value = _extract_numeric(
                    payload,
                    ("current", "current_a"),
                    fuzzy=("current", "amp", "irms"),
                )
                if current_value is not None:
                    current_samples.append((entry, current_value))

                voltage_value = _extract_numeric(
                    payload,
                    ("voltage", "voltage_v"),
                    fuzzy=("voltage", "volt", "vrms"),
                )
                if voltage_value is not None:
                    voltage_samples.append((entry, voltage_value))

            if not pf_samples:
                continue

            min_entry, min_pf = min(pf_samples, key=lambda pair: pair[1])
            max_entry, max_pf = max(pf_samples, key=lambda pair: pair[1])
            avg_pf = sum(pf for _, pf in pf_samples) / len(pf_samples)

            min_current_entry = max_current_entry = None
            min_current = max_current = avg_current = 0.0
            if current_samples:
                min_current_entry, min_current = min(current_samples, key=lambda pair: pair[1])
                max_current_entry, max_current = max(current_samples, key=lambda pair: pair[1])
                avg_current = sum(val for _, val in current_samples) / len(current_samples)

            min_voltage_entry = max_voltage_entry = None
            min_voltage = max_voltage = avg_voltage = 0.0
            if voltage_samples:
                min_voltage_entry, min_voltage = min(voltage_samples, key=lambda pair: pair[1])
                max_voltage_entry, max_voltage = max(voltage_samples, key=lambda pair: pair[1])
                avg_voltage = sum(val for _, val in voltage_samples) / len(voltage_samples)

            start_kwh = end_kwh = None
            for _, payload in samples:
                kwh_value = _extract_numeric(
                    payload,
                    ("kwh",),
                    fuzzy=("kwh", "energy", "wh"),
                )
                if kwh_value is None:
                    continue
                if start_kwh is None:
                    start_kwh = kwh_value
                end_kwh = kwh_value
            total_kwh = 0.0
            if start_kwh is not None and end_kwh is not None:
                total_kwh = max(0.0, end_kwh - start_kwh)

            defaults = {
                'min_power_factor': min_pf,
                'max_power_factor': max_pf,
                'min_power_factor_time': min_entry.timestamp,
                'max_power_factor_time': max_entry.timestamp,
                'avg_power_factor': avg_pf,
                'total_kwh': total_kwh,
                'min_current': min_current,
                'max_current': max_current,
                'avg_current': avg_current,
                'min_current_time': min_current_entry.timestamp if min_current_entry else None,
                'max_current_time': max_current_entry.timestamp if max_current_entry else None,
                'min_voltage': min_voltage,
                'max_voltage': max_voltage,
                'avg_voltage': avg_voltage,
                'min_voltage_time': min_voltage_entry.timestamp if min_voltage_entry else None,
                'max_voltage_time': max_voltage_entry.timestamp if max_voltage_entry else None,
                'data_points': len(pf_samples),
            }

            _, _ = ShiftReport.objects.update_or_create(
                shift=shift,
                device=device,
                date=target_date,
                defaults=defaults,
            )
            generated_count += 1

    return generated_count

@login_required
def device_list(request):
    owned_ids = set(
        Device.objects.filter(
            device_owner=request.user,
            provisioning_state=Device.ProvisioningState.ACTIVE,
        ).values_list('id', flat=True)
    )
    devices = (
        Device.objects.filter(
            Q(device_owner=request.user) | Q(shared_with=request.user),
            provisioning_state=Device.ProvisioningState.ACTIVE,
        )
        .select_related('device_owner')
        .prefetch_related('device_shares__user')
        .distinct()
    )
    current_time = timezone.now()
    
    # Queue polling jobs for devices that are due
    for device in devices:
        if device.id not in owned_ids:
            continue  # Shared devices are read-only for this user
        try:
            if device.last_updated is None or (
                (current_time - device.last_updated).total_seconds() > device.polling_interval
            ):
                poll_device_task.delay(device.id)
        except (CeleryError, OperationalError) as exc:
            logger.warning("Unable to queue poll for device %s: %s", device.id, exc)
            success, payload = device.poll_device()
            if not success:
                logger.warning("Fallback poll failed for device %s: %s", device.id, payload)
        except Exception as exc:  # pragma: no cover - defensive guard
            logger.exception("Unexpected error scheduling poll for device %s", device.id)
    
    return render(
        request,
        'devices/device_list.html',
        {
            'devices': devices,
            'owned_device_ids': list(owned_ids),
        },
    )


@login_required
@require_POST
def download_devices_csv(request):
    password = request.POST.get('password', '')
    redirect_url = request.POST.get('next')

    if not request.user.check_password(password):
        messages.error(request, 'Password verification failed. Please try again.')
        if redirect_url:
            return redirect(redirect_url)
        return redirect('dashboard')

    devices = Device.objects.filter(device_owner=request.user).order_by('created_at')

    response = HttpResponse(content_type='text/csv')
    timestamp = timezone.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="devices-{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Device',
        'Installed At',
        'Address',
        'Address Type',
        'Status',
        'Created At',
        'Last Updated',
        'Polling Interval (seconds)',
    ])

    for device in devices:
        created_at = (
            timezone.localtime(device.created_at).strftime('%Y-%m-%d %H:%M:%S')
            if device.created_at
            else ''
        )
        last_updated = (
            timezone.localtime(device.last_updated).strftime('%Y-%m-%d %H:%M:%S')
            if device.last_updated
            else ''
        )
        writer.writerow([
            device.get_device_type_display() if hasattr(device, 'get_device_type_display') else device.device_type,
            device.located_at or '',
            device.device_address or '',
            device.get_address_type_display() if hasattr(device, 'get_address_type_display') else (device.address_type or ''),
            'Active' if device.is_active else 'Inactive',
            created_at,
            last_updated,
            device.polling_interval,
        ])

    return response


@login_required
@require_POST
def device_config_detail(request, device_id):
    device = get_object_or_404(
        Device.objects.prefetch_related('device_shares__user'), id=device_id
    )
    role = _get_device_role(device, request.user)
    if role not in {
        DeviceShare.AccessLevel.INSPECTOR,
        DeviceShare.AccessLevel.MANAGER,
    }:
        return JsonResponse({'error': 'Permission denied.'}, status=403)
    password = request.POST.get('password', '')

    if not request.user.check_password(password):
        return JsonResponse({'error': 'Password verification failed.'}, status=403)

    def format_timestamp(value):
        if not value:
            return ''
        localized = timezone.localtime(value)
        return localized.strftime('%Y-%m-%d %H:%M:%S')

    payload = {
        'id': device.id,
        'name': device.get_device_type_display() if hasattr(device, 'get_device_type_display') else device.device_type,
        'location': device.located_at or 'N/A',
        'address': device.device_address or 'N/A',
        'status': 'Active' if device.is_active else 'Inactive',
        'created': format_timestamp(device.created_at),
        'updated': format_timestamp(device.last_updated),
    }

    return JsonResponse({'device': payload})

@login_required
def add_device(request):
    if request.method == 'POST':
        Device.purge_expired_pending(owner=request.user)
        form = DeviceForm(request.POST)
        if form.is_valid():
            device = form.save(commit=False)
            device.device_owner = request.user
            device.is_active = False
            device.provisioning_state = Device.ProvisioningState.PENDING

            existing = Device.objects.filter(
                device_owner=request.user,
                device_address=device.device_address,
            ).first()

            if existing:
                if existing.provisioning_state == Device.ProvisioningState.PENDING:
                    if existing.has_active_provisioning_window():
                        form.add_error(
                            'device_address',
                            'A device at this address is already pending verification. Use the existing provisioning token or wait for it to expire.',
                        )
                        return render(request, 'devices/add_device.html', {'form': form})
                    existing.delete()
                else:
                    form.add_error(
                        'device_address',
                        'You already have an active device registered with this address.',
                    )
                    return render(request, 'devices/add_device.html', {'form': form})

            try:
                with transaction.atomic():
                    device.save()
            except IntegrityError:
                logger.info(
                    "Duplicate device registration prevented for user=%s address=%s",
                    request.user.id,
                    device.device_address,
                )
                form.add_error(
                    'device_address',
                    'A device with this address already exists. Please verify the address or remove the existing device.',
                )
                return render(request, 'devices/add_device.html', {'form': form})

            token, token_obj = DeviceProvisioningToken.issue(
                device,
                created_by=request.user,
                metadata={'channel': 'ui', 'initiated_during_create': True},
            )
            request.session['provision_token_value'] = token
            request.session['provision_token_expiry'] = (
                token_obj.expires_at.isoformat() if token_obj.expires_at else ''
            )
            request.session['provision_token_device'] = device.id
            messages.success(
                request,
                'Device saved in pending state. Share the provisioning token to complete setup.',
            )
            return redirect('devices:device_provisioning', device_id=device.id)
    else:
        form = DeviceForm()

    return render(request, 'devices/add_device.html', {'form': form})

@login_required
def device_detail(request, device_id):
    device = get_object_or_404(
        Device.objects.prefetch_related('device_shares__user'), id=device_id
    )
    is_owner = device.device_owner_id == request.user.id

    if device.provisioning_state != Device.ProvisioningState.ACTIVE and not is_owner:
        raise Http404("Device not found")
    if device.provisioning_state == Device.ProvisioningState.PENDING and is_owner:
        messages.info(
            request,
            'This device is awaiting provisioning. Share the token to complete setup.',
        )

    user_role = _get_device_role(device, request.user)
    if user_role is None:
        raise Http404("Device not found")

    can_manage = user_role == DeviceShare.AccessLevel.MANAGER
    can_view_config = user_role in (
        DeviceShare.AccessLevel.INSPECTOR,
        DeviceShare.AccessLevel.MANAGER,
    )
    is_andon = device.device_type == 'andon'
    andon_snapshot = None
    if is_andon:
        station = _find_andon_station(device)
        if station:
            andon_snapshot = _build_andon_snapshot(station)

    current_time = timezone.now()

    accessible_devices = (
        Device.objects.filter(
            Q(device_owner=request.user) | Q(shared_with=request.user),
            provisioning_state=Device.ProvisioningState.ACTIVE,
        )
        .prefetch_related('device_shares__user')
        .distinct()
        .order_by('id')
    )
    next_device = accessible_devices.filter(id__gt=device_id).order_by('id').first()
    prev_device = accessible_devices.filter(id__lt=device_id).order_by('-id').first()

    if request.GET.get('poll') == 'true':
        if device.provisioning_state != Device.ProvisioningState.ACTIVE:
            messages.error(request, 'Device must complete provisioning before polling.')
            return redirect('devices:device_provisioning', device_id=device_id)
        if not can_manage:
            messages.error(request, 'You do not have permission to trigger polling.')
            return redirect('devices:device_detail', device_id=device_id)

        logger.debug("Manual poll requested for device %s", device_id)
        try:
            poll_device_task.delay(device.id, force=True)
            messages.success(request, 'Polling requested. Latest data will appear shortly.')
        except (CeleryError, OperationalError) as exc:
            logger.warning("Unable to queue manual poll for device %s: %s", device.id, exc)
            success, data = device.poll_device()
            if success:
                messages.success(request, 'Polled device immediately (Celery unavailable).')
            else:
                messages.error(request, f'Failed to update device data: {data}')
            return redirect('devices:device_detail', device_id=device_id)
        except Exception as exc:  # pragma: no cover - defensive guard
            logger.exception("Unexpected error queuing poll for device %s", device.id)
            messages.error(request, f'Unexpected error queuing poll: {exc}')
        return redirect('devices:device_detail', device_id=device_id)

    if can_manage and device.provisioning_state == Device.ProvisioningState.ACTIVE:
        try:
            should_poll = (
                device.last_updated is None or
                (current_time - device.last_updated).total_seconds() > device.polling_interval
            )

            if should_poll:
                logger.debug(
                    "Auto-queueing poll for device %s (last update: %s)",
                    device_id,
                    device.last_updated,
                )
                try:
                    poll_device_task.delay(device.id)
                except (CeleryError, OperationalError) as exc:
                    logger.warning(
                        "Unable to queue auto poll for device %s, falling back: %s",
                        device.id,
                        exc,
                    )
                    success, data = device.poll_device()
                    if not success:
                        messages.warning(request, f'Auto-poll failed: {data}')
        except CeleryError as exc:
            messages.warning(request, f'Unable to schedule auto-update: {exc}')
        except Exception as exc:  # pragma: no cover - defensive guard
            logger.exception("Error during auto-poll scheduling for device %s", device.id)
            messages.warning(request, f'Error during auto-update: {exc}')

    historical_data = device.historical_data.all()[:100]  # Last 100 readings
    
    # Prepare data for charts
    timestamps = []
    voltage_data = []
    current_data = []
    power_factor_data = []
    kwh_data = []
    
    for entry in reversed(historical_data):
        if isinstance(entry.value, dict):
            timestamps.append(entry.timestamp.strftime('%H:%M:%S'))
            voltage_data.append(entry.value.get('voltage', 0))
            current_data.append(entry.value.get('current', 0))
            power_factor_data.append(entry.value.get('power_factor', 0))
            kwh_data.append(entry.value.get('kwh', 0))
    
    chart_data = {
        'timestamps': json.dumps(timestamps),
        'voltage': json.dumps(voltage_data),
        'current': json.dumps(current_data),
        'power_factor': json.dumps(power_factor_data),
        'kwh': json.dumps(kwh_data)
    }
    
    shared_entries = list(
        device.device_shares.select_related('user').order_by('user__username')
    )

    return render(
        request,
        'devices/device_detail.html',
        {
            'device': device,
            'historical_data': historical_data,
            'chart_data': chart_data,
            'next_device': next_device,
            'prev_device': prev_device,
            'is_owner': is_owner,
            'access_role': user_role,
            'can_view_config': can_view_config,
            'can_manage_device': can_manage,
            'is_andon': is_andon,
            'andon_snapshot': andon_snapshot,
            'config_url': reverse('devices:device_config_detail', args=[device.id]),
            'shared_entries': shared_entries,
        },
    )

@login_required
def remove_device(request, device_id):
    device = get_object_or_404(
        Device.objects.prefetch_related('device_shares__user'), id=device_id
    )
    role = _get_device_role(device, request.user)
    if role != DeviceShare.AccessLevel.MANAGER:
        raise Http404("Device not found")
    if request.method == 'POST':
        device.delete()
        messages.success(request, f'Device "{device.device_type} at {device.located_at}" has been removed.')
        return redirect('devices:device_list')
    return render(request, 'devices/remove_device.html', {'device': device})



@login_required
def device_provisioning(request, device_id):
    device = get_object_or_404(Device, id=device_id, device_owner=request.user)
    issued_token = None
    issued_expires_at = None
    default_lifetime = DeviceProvisioningToken.DEFAULT_LIFETIME
    default_minutes = (
        int(default_lifetime.total_seconds() // 60)
        if default_lifetime
        else ''
    )

    session_device_id = request.session.pop('provision_token_device', None)
    session_token = request.session.pop('provision_token_value', None)
    session_expiry = request.session.pop('provision_token_expiry', None)
    if session_token and session_device_id == device.id:
        issued_token = session_token
        if session_expiry:
            parsed = parse_datetime(session_expiry)
            if parsed and timezone.is_aware(parsed):
                issued_expires_at = timezone.localtime(parsed)
            else:
                issued_expires_at = parsed
    else:
        if session_device_id is not None and session_device_id != device.id:
            request.session['provision_token_device'] = session_device_id
        if session_token and session_device_id != device.id:
            request.session['provision_token_value'] = session_token
        if session_expiry and session_device_id != device.id:
            request.session['provision_token_expiry'] = session_expiry

    if request.method == 'POST':
        action = request.POST.get('action', 'issue')
        if action == 'revoke':
            password = (request.POST.get('confirm_password') or '').strip()
            if not password:
                messages.error(request, 'Enter your account password to revoke the credential.')
                return redirect('devices:device_provisioning', device_id=device.id)
            if not request.user.check_password(password):
                messages.error(request, 'Incorrect password. Credential was not revoked.')
                return redirect('devices:device_provisioning', device_id=device.id)

            device.clear_api_secret()
            messages.success(request, 'Device API credential revoked.')
            return redirect('devices:device_provisioning', device_id=device.id)

        lifetime_raw = (request.POST.get('lifetime_minutes') or '').strip()
        lifetime = None
        if lifetime_raw:
            try:
                minutes = max(1, int(lifetime_raw))
            except (TypeError, ValueError):
                messages.error(request, 'Token lifetime must be a whole number of minutes.')
                return redirect('devices:device_provisioning', device_id=device.id)
            lifetime = timedelta(minutes=minutes)

        notes = (request.POST.get('notes') or '').strip()
        metadata = {
            'notes': notes or None,
            'issued_from_ip': request.META.get('REMOTE_ADDR'),
            'channel': 'ui',
        }
        token, token_obj = DeviceProvisioningToken.issue(
            device,
            created_by=request.user,
            lifetime=lifetime,
            metadata={k: v for k, v in metadata.items() if v},
        )
        issued_token = token
        if token_obj.expires_at:
            issued_expires_at = timezone.localtime(token_obj.expires_at)
        else:
            issued_expires_at = None
        messages.success(request, 'Provisioning token generated successfully.')

    recent_queryset = device.provisioning_tokens.select_related('created_by').order_by('-created_at')[:10]
    now_value = timezone.now()
    recent_tokens = []
    for token in recent_queryset:
        if token.used_at:
            status = 'claimed'
        elif token.expires_at and token.expires_at <= now_value:
            status = 'expired'
        else:
            status = 'pending'
        recent_tokens.append({'token': token, 'status': status})

    claim_url = request.build_absolute_uri(reverse('api:device-claim'))
    ingest_url = request.build_absolute_uri(reverse('api:device-data-ingest'))

    context = {
        'device': device,
        'issued_token': issued_token,
        'issued_expires_at': issued_expires_at,
        'default_minutes': default_minutes,
        'recent_tokens': recent_tokens,
        'has_active_secret': bool(device.device_secret_hash),
        'provisioning_state': device.provisioning_state,
        'claim_url': claim_url,
        'ingest_url': ingest_url,
    }
    return render(request, 'devices/device_provision.html', context)


@login_required
def manage_shifts(request):
    shifts = Shift.objects.all()
    return render(request, 'devices/manage_shifts.html', {'shifts': shifts})

@login_required
def add_shift(request):
    if request.method == 'POST':
        try:
            shift = Shift(
                name=request.POST['name'],
                start_time=request.POST['start_time'],
                end_time=request.POST['end_time']
            )
            shift.full_clean()  # Run validation
            shift.save()
            messages.success(request, f'Shift "{shift.name}" added successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error adding shift: {str(e)}')
    return redirect('devices:manage_shifts')

@login_required
def edit_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)
    if request.method == 'POST':
        try:
            shift.name = request.POST['name']
            shift.start_time = request.POST['start_time']
            shift.end_time = request.POST['end_time']
            shift.full_clean()  # Run validation
            shift.save()
            messages.success(request, f'Shift "{shift.name}" updated successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error updating shift: {str(e)}')
    return redirect('devices:manage_shifts')

@login_required
def toggle_shift(request, shift_id):
    if request.method == 'POST':
        shift = get_object_or_404(Shift, id=shift_id)
        shift.is_active = not shift.is_active
        shift.save()
        return JsonResponse({'status': 'success'})

@login_required
def shift_reports(request):
    # Handle manual report generation
    if request.method == 'POST' and request.POST.get('generate_reports'):
        generate_date_str = request.POST.get('generate_date')
        if generate_date_str:
            try:
                generate_date = datetime.strptime(generate_date_str, '%Y-%m-%d').date()
                generated_count = generate_shift_reports_for_date(generate_date)
                messages.success(request, f'Generated {generated_count} shift reports for {generate_date}')
            except ValueError:
                messages.error(request, 'Invalid date format')
        else:
            messages.error(request, 'Please select a date to generate reports')

    shifts = Shift.objects.all()
    reports = ShiftReport.objects.select_related('shift', 'device').order_by('-date', 'shift__start_time')

    # Filter by date if provided
    date_str = request.GET.get('date')
    if date_str:
        try:
            filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            reports = reports.filter(date=filter_date)
        except ValueError:
            messages.error(request, 'Invalid date format')

    return render(request, 'devices/shift_reports.html', {
        'shifts': shifts,
        'reports': reports,
        'selected_date': date_str
    })


@login_required
def bulk_share_devices(request):
    owned_devices = (
        Device.objects.filter(device_owner=request.user)
        .select_related('device_owner')
        .prefetch_related('device_shares__user')
        .order_by('created_at')
    )

    identifier = ''
    selected_ids = []
    available_users = (
        User.objects.filter(is_active=True)
        .exclude(id=request.user.id)
        .order_by('username')
    )
    selected_role = DeviceShare.default_role()
    role_choices = DeviceShare.role_options()
    role_lookup = {choice['value']: choice for choice in role_choices}

    if request.method == 'POST':
        identifier = (request.POST.get('identifier') or '').strip()
        selected_ids = request.POST.getlist('device_ids')
        selected_role = request.POST.get('role') or DeviceShare.default_role()
        action = request.POST.get('action')
        password = request.POST.get('password') or ''
        remove_all = request.POST.get('remove_all') == '1'
        remove_user_ids = request.POST.getlist('remove_user_ids')

        if action == 'share' and selected_role not in role_lookup:
            messages.error(request, 'Select a valid access level.')
        elif not action:
            messages.error(request, 'Select an action to perform.')
        elif not selected_ids:
            messages.error(request, 'Select at least one device first.')
        else:
            devices_to_share = owned_devices.filter(id__in=selected_ids)

            if not devices_to_share.exists():
                messages.error(request, 'No matching devices were found.')
            elif action == 'share':
                if not identifier:
                    messages.error(request, 'Enter an email address or username to share with.')
                else:
                    user_to_share = (
                        User.objects.filter(
                            Q(email__iexact=identifier) | Q(username__iexact=identifier)
                        )
                        .exclude(id=request.user.id)
                        .first()
                    )

                    if not user_to_share:
                        messages.error(request, f'No user found with email or username "{identifier}".')
                    elif not password:
                        messages.error(request, 'Enter your password to share devices.')
                    elif not request.user.check_password(password):
                        messages.error(request, 'Password verification failed.')
                    else:
                        shared_count = 0
                        updated_count = 0
                        already_shared = 0
                        for device in devices_to_share:
                            existing_share = DeviceShare.objects.filter(
                                device=device,
                                user=user_to_share,
                            ).first()
                            if existing_share:
                                if existing_share.role == selected_role:
                                    already_shared += 1
                                else:
                                    existing_share.role = selected_role
                                    existing_share.save(update_fields=['role'])
                                    updated_count += 1
                            else:
                                DeviceShare.objects.create(
                                    device=device,
                                    user=user_to_share,
                                    role=selected_role,
                                )
                                shared_count += 1

                        display_name = user_to_share.get_short_name() or user_to_share.username
                        role_label = role_lookup.get(selected_role, {}).get('label', 'access')
                        if shared_count:
                            messages.success(
                                request,
                                f'Granted {role_label.lower()} to {display_name} on {shared_count} device(s).',
                            )
                        if updated_count:
                            messages.info(
                                request,
                                f'Updated access level for {display_name} on {updated_count} device(s).',
                            )
                        if already_shared and not updated_count:
                            messages.info(
                                request,
                                f'{display_name} already had the selected access on {already_shared} device(s).',
                            )
                        return redirect('devices:bulk_share_devices')
            elif action == 'remove':
                if remove_all:
                    removed_total = 0
                    cleared_devices = 0
                    untouched_devices = 0
                    for device in devices_to_share:
                        shares_qs = device.device_shares.all()
                        share_count = shares_qs.count()
                        if share_count:
                            removed_total += share_count
                            cleared_devices += 1
                            shares_qs.delete()
                        else:
                            untouched_devices += 1

                    if removed_total:
                        messages.success(
                            request,
                            f'Removed all shared users ({removed_total} assignment(s)) from {cleared_devices} device(s).',
                        )
                    if untouched_devices:
                        messages.info(
                            request,
                            f'{untouched_devices} device(s) already had no shared users.',
                        )
                    return redirect('devices:bulk_share_devices')

                elif remove_user_ids:
                    shares_to_remove = DeviceShare.objects.filter(
                        device__in=devices_to_share,
                        user_id__in=remove_user_ids,
                    )

                    if not shares_to_remove.exists():
                        messages.error(request, 'No matching users selected for removal.')
                    else:
                        removed_total = shares_to_remove.count()
                        affected_devices = shares_to_remove.values('device_id').distinct().count()
                        shares_to_remove.delete()
                        messages.success(
                            request,
                            f'Removed {removed_total} sharing assignment(s) across {affected_devices} device(s).',
                        )
                        return redirect('devices:bulk_share_devices')

                elif identifier:
                    user_to_share = (
                        User.objects.filter(
                            Q(email__iexact=identifier) | Q(username__iexact=identifier)
                        )
                        .exclude(id=request.user.id)
                        .first()
                    )

                    if not user_to_share:
                        messages.error(request, f'No user found with email or username "{identifier}".')
                    else:
                        removed_count = 0
                        not_shared = 0
                        for device in devices_to_share:
                            deleted, _ = DeviceShare.objects.filter(
                                device=device,
                                user=user_to_share,
                            ).delete()
                            if deleted:
                                removed_count += 1
                            else:
                                not_shared += 1

                        display_name = user_to_share.get_short_name() or user_to_share.username
                        if removed_count:
                            messages.success(
                                request,
                                f'Removed {display_name} from {removed_count} device(s).',
                            )
                        if not_shared:
                            messages.info(
                                request,
                                f'{display_name} did not have access to {not_shared} selected device(s).',
                            )
                        return redirect('devices:bulk_share_devices')
                else:
                    messages.error(request, 'Select at least one user to remove.')
            else:
                messages.error(request, 'Unknown action requested.')

    context = {
        'devices': owned_devices,
        'identifier': identifier,
        'available_users': available_users,
        'selected_ids': selected_ids,
        'role_choices': role_choices,
        'selected_role': selected_role,
    }
    return render(request, 'devices/bulk_share_devices.html', context)


@login_required
def device_report(request, device_id):
    device = get_object_or_404(Device, id=device_id, device_owner=request.user)
    period = request.GET.get('period', 'daily')  # Options: daily, weekly, monthly
    format_type = request.GET.get('format', 'csv')  # Options: csv, excel
    
    # Calculate date range
    end_date = timezone.now()
    if period == 'daily':
        start_date = end_date - timedelta(days=1)
    elif period == 'weekly':
        start_date = end_date - timedelta(weeks=1)
    else:  # monthly
        start_date = end_date - timedelta(days=30)
        
    # Get regular data and shift reports for the period
    data = device.historical_data.filter(
        timestamp__range=(start_date, end_date)
    ).order_by('timestamp')
    
    shift_reports = device.shift_reports.filter(
        date__range=(start_date.date(), end_date.date())
    ).select_related('shift')
    
    # Prepare the response
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{device.device_type}_{device.located_at}_{period}_report.csv"'
        
        writer = csv.writer(response)
        
        # Write shift summary section
        writer.writerow(['Shift Reports'])
        writer.writerow(['Date', 'Shift', 'Min PF', 'Min PF Time', 'Max PF', 'Max PF Time', 'Avg PF', 'Total kWh'])
        for report in shift_reports:
            writer.writerow([
                report.date,
                report.shift.name,
                f"{report.min_power_factor:.3f}",
                report.min_power_factor_time.strftime('%Y-%m-%d %H:%M:%S'),
                f"{report.max_power_factor:.3f}",
                report.max_power_factor_time.strftime('%Y-%m-%d %H:%M:%S'),
                f"{report.avg_power_factor:.3f}",
                f"{report.total_kwh:.1f}"
            ])
        
        # Add a blank line between sections
        writer.writerow([])
        
        # Write detailed readings section
        writer.writerow(['Detailed Readings'])
        writer.writerow(['Timestamp', 'Voltage (V)', 'Current (A)', 'Power Factor', 'kWh', 'kVAh'])
        for entry in data:
            writer.writerow([
                entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                entry.value.get('voltage', ''),
                entry.value.get('current', ''),
                entry.value.get('power_factor', ''),
                entry.value.get('kwh', ''),
                entry.value.get('kwah', '')
            ])
            
        return response
    else:
        # For Excel format, we need to install openpyxl
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{device.device_type} Report"

            current_row = 1

            # Shift Reports Section
            if shift_reports.exists():
                ws[f"A{current_row}"] = "Shift Reports"
                current_row += 1

                # Shift headers
                shift_headers = ['Date', 'Shift', 'Min PF', 'Min PF Time', 'Max PF', 'Max PF Time', 'Avg PF', 'Total kWh']
                for col, header in enumerate(shift_headers, 1):
                    ws[f"{get_column_letter(col)}{current_row}"] = header
                current_row += 1

                # Shift data
                for report in shift_reports:
                    ws[f"A{current_row}"] = report.date.strftime('%Y-%m-%d')
                    ws[f"B{current_row}"] = report.shift.name
                    ws[f"C{current_row}"] = f"{report.min_power_factor:.3f}"
                    ws[f"D{current_row}"] = report.min_power_factor_time.strftime('%Y-%m-%d %H:%M:%S')
                    ws[f"E{current_row}"] = f"{report.max_power_factor:.3f}"
                    ws[f"F{current_row}"] = report.max_power_factor_time.strftime('%Y-%m-%d %H:%M:%S')
                    ws[f"G{current_row}"] = f"{report.avg_power_factor:.3f}"
                    ws[f"H{current_row}"] = f"{report.total_kwh:.1f}"
                    current_row += 1

                # Add blank row
                current_row += 1

            # Detailed Readings Section
            ws[f"A{current_row}"] = "Detailed Readings"
            current_row += 1

            # Headers
            headers = ['Timestamp', 'Voltage (V)', 'Current (A)', 'Power Factor', 'kWh', 'kVAh']
            for col, header in enumerate(headers, 1):
                ws[f"{get_column_letter(col)}{current_row}"] = header
            current_row += 1

            # Data
            for entry in data:
                ws[f"A{current_row}"] = entry.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ws[f"B{current_row}"] = entry.value.get('voltage', '')
                ws[f"C{current_row}"] = entry.value.get('current', '')
                ws[f"D{current_row}"] = entry.value.get('power_factor', '')
                ws[f"E{current_row}"] = entry.value.get('kwh', '')
                ws[f"F{current_row}"] = entry.value.get('kwah', '')
                current_row += 1
            
            # Auto-adjust column widths
            max_cols = max(len(headers), 8 if shift_reports.exists() else len(headers))
            for col in range(1, max_cols + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{device.device_type}_{device.located_at}_{period}_report.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl package. Using CSV format instead.')
            return redirect(request.path + '?period=' + period + '&format=csv')
@login_required
def share_device(request, device_id):
    device = get_object_or_404(Device, id=device_id, device_owner=request.user)

    shared_entries = list(
        device.device_shares.select_related('user').order_by('user__username')
    )
    shared_user_ids = [entry.user_id for entry in shared_entries]
    available_users = (
        User.objects.filter(is_active=True)
        .exclude(id__in=[request.user.id, *shared_user_ids])
        .order_by('username')
    )
    role_choices = DeviceShare.role_options()
    role_lookup = {choice['value']: choice for choice in role_choices}
    selected_role = DeviceShare.default_role()

    if request.method == 'POST':
        identifier = (request.POST.get('identifier') or request.POST.get('username') or '').strip()
        selected_role = request.POST.get('role') or DeviceShare.default_role()

        if selected_role not in role_lookup:
            messages.error(request, 'Select a valid access level.')
        elif not identifier:
            messages.error(request, 'Please enter an email address or username to share with.')
        else:
            user_to_share = (
                User.objects.filter(
                    Q(email__iexact=identifier) | Q(username__iexact=identifier)
                )
                .exclude(id=request.user.id)
                .first()
            )

            if not user_to_share:
                messages.error(request, f'No user found with email or username "{identifier}".')
            else:
                share, created = DeviceShare.objects.get_or_create(
                    device=device,
                    user=user_to_share,
                    defaults={'role': selected_role},
                )
                display_name = user_to_share.get_short_name() or user_to_share.username
                if not created and share.role == selected_role:
                    messages.info(request, f'{display_name} already has this access level.')
                else:
                    if not created:
                        share.role = selected_role
                        share.save(update_fields=['role'])
                        messages.success(request, f'Updated access for {display_name}.')
                    else:
                        messages.success(request, f'Shared {device} with {display_name}.')
                return redirect('devices:share_device', device_id=device.id)

        return redirect('devices:share_device', device_id=device.id)

    context = {
        'device': device,
        'available_users': available_users,
        'shared_entries': shared_entries,
        'role_choices': role_choices,
        'selected_role': selected_role,
    }
    return render(request, 'devices/share_device.html', context)


@login_required
def remove_shared_user(request, device_id, user_id):
    device = get_object_or_404(Device, id=device_id, device_owner=request.user)

    if request.method == 'POST':
        deleted, _ = DeviceShare.objects.filter(device=device, user_id=user_id).delete()
        if deleted:
            messages.success(request, 'Removed sharing access.')
        else:
            messages.info(request, 'User did not have access to this device.')

    return redirect('devices:share_device', device_id=device.id)
